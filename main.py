# Import the required modules
import openpyxl
import openpyxl.cell.cell as Cell
import pandas as pd
import streamlit as st
import shutil, os, tempfile
from datetime import datetime as dt
from functions.Extract_params import *
from functions.From_q import *
from functions.weather_stations import *
from functions.aia_api import find_selected_indices, build_aia_payload, call_aia_api
import json

_CROP_NAME = object()  # sentinel: write crop.capitalize() into this column


def write_products_to_ws(ws, crops, products_by_crop, col_map):
    row = 2
    for i, crop in enumerate(crops):
        if i > 0:
            row += len(products_by_crop[crops[i - 1]])
        for space, product in enumerate(products_by_crop[crop]):
            for col, key in col_map.items():
                ws.cell(row + space, col).value = (
                    crop.capitalize() if key is _CROP_NAME else product[key]
                )


STATE_REMAP = {"nw_western_australia": "wa_nw", "sw_western_australia": "wa_sw"}


# Get current path
cwd = os.getcwd()

st.title("Carbon accounting tool")

tool = st.sidebar.radio("Select which tools you want to run:", ["Extraction", "API"])

if tool == "Extraction":
    st.header("Questionnaire extraction")

    zipfiles = st.file_uploader(
        "Upload your questionnaire form as a csv format:",
        "zip",
        accept_multiple_files=True,
    )

    try:
        crops, crop_specific_input, questionnaire_df, veg_df = from_the_top(zipfiles)

        cols_to_drop = [
            "ObjectID",
            "GlobalID",
            "CreationDate",
            "Creator",
            "EditDate",
            "Editor",
            "x",
            "y",
        ]

        questionnaire_df = questionnaire_df.drop(cols_to_drop, axis=1)

        try:
            production_year = dt.strptime(
                questionnaire_df["production_year"].iloc[0], "%d/%m/%Y %I:%M:%S %p"
            ).year
        except TypeError:
            production_year = dt(questionnaire_df["production_year"].iloc[0], 1, 1).year

    except AttributeError:
        st.write("Haven't uploaded a zip of Survey123 production data!")
    except UnboundLocalError:
        st.write("Haven't uploaded a zip of Survey123 production data!")

    planting_shapes = st.file_uploader(
        "Upload your planting shapefile (zip or all of it)",
        accept_multiple_files=True,
        key="PlantingShape",
    )

    # Number of crop in the questionnaire
    if st.button("Get your crop types", "CropType"):
        # Read in the form as csv
        st.write(crops)

    tab1, tab2, tab3 = st.tabs(
        ["Check questionnaire", "Check fert/chem input", "Get weather from DPIRD's API"]
    )

    with tab1:
        try:
            df_t = {}
            for label, content in questionnaire_df.items():
                df_t[label] = content.iloc[0]
            # Transform the questionnaire into a df for first pass
            st.dataframe(df_t)
        except NameError:
            st.write("Nothing to see here!")

    with tab2:
        try:
            crop = st.radio("Choose the crop to view", crops)
            input = st.radio(
                "Choose an input to review",
                ["fert", "fungicide", "herbicide", "insecticide", "chem_other"],
            )
            st.dataframe(crop_specific_input[crop][input], hide_index=True)
        except NameError:
            st.markdown(":woman-gesturing-no:!:man-gesturing-no:!\
                        Quickly upload files to discover the secret lies here")

    with tab3:
        # Upload shapefile for SILO's API (weather data)
        shapes = st.file_uploader(
            "Upload all of your shapefile for weather data or the compressed file:",
            accept_multiple_files=True,
        )

        # get API data
        api_url = "https://www.longpaddock.qld.gov.au/cgi-bin/silo"

        params = {
            "format": "near",
            "station": 10619,  # Nyabing weather station
            "radius": 800,  # in km
        }
        url = api_url + "/PatchedPointDataset.php?" + urllib.parse.urlencode(params)

        with urllib.request.urlopen(url) as remote:
            data = remote.read()

        # write API data as file and then as Pd DF
        with io.BytesIO() as f:
            f.write(data)
            f.seek(0)
            weather_stations = pd.read_csv(f, delimiter="|")

        try:  # Incase there are no files (don't want to scare people away)
            # Get the coordinate from the shapefile

            gdf = read_shapes(shapes)

            centroid = gdf.dissolve().centroid
            lon = centroid.x[0]
            lat = centroid.y[0]

            # A df of nearest station
            nearest_station = get_nearby_stations(lat, lon, weather_stations)

            # To show four nearest weather station with the
            # fraction of data from BOM
            st.write(nearest_station)

            endYear = int(
                st.text_input("Input the end year (YYYY):", dt.now().year - 1)
            )

            # Get a list of weather data from all four weather station
            weather_dfs = to_list_dfs(endYear, nearest_station)

            # Choose the data from a weather station or
            # a weighted average of multiple stations
            selected_stations = st.multiselect(
                "Select your weather station (one or multiples):",
                nearest_station.iloc[:, 0].to_list(),
            )
            # Indexes to go through the list of selected
            # station and list of all weather station's df
            i, j = 0, 0
            # If multiples stations are selected create a
            # list
            if len(selected_stations) > 1:
                extracted_df = []
            if len(selected_stations) == 0:
                raise Exception("Haven't selected a weather station")

            # Append the selected stations into the extracted list
            while i < len(selected_stations) and j < len(weather_dfs):
                if weather_dfs[j].iloc[0, 0] == selected_stations[i]:
                    try:
                        extracted_df.append(weather_dfs[j])
                    except NameError:
                        extracted_df = weather_dfs[j]
                    j = 0
                    i += 1
                else:
                    j += 1

            # Create an empty df for extracted data or
            # weighted average if multiple stations
            daily_df = pd.DataFrame()
            if isinstance(extracted_df, list):
                daily_df["Date"] = extracted_df[0]["YYYY-MM-DD"]
            else:
                daily_df["Date"] = extracted_df["YYYY-MM-DD"]
            daily_df["Year"] = [int(i[0:4]) for i in daily_df["Date"]]
            daily_df["Rain"] = weighted_ave_col(
                extracted_df, "daily_rain", nearest_station, selected_stations
            )
            daily_df["ETShortCrop"] = weighted_ave_col(
                extracted_df, "et_short_crop", nearest_station, selected_stations
            )
            daily_df["ETTallCrop"] = weighted_ave_col(
                extracted_df, "et_tall_crop", nearest_station, selected_stations
            )

            rain, eto_short, eto_tall = annual_summary(daily_df, endYear)
            rain_long, eto_short_long, eto_tall_long = longTerms_summary(daily_df)

            # Create a folder for saving and linkage
            # to the excel writing
            with tempfile.TemporaryDirectory() as td:
                daily_df.to_csv(
                    os.path.join(
                        td,
                        f"{'+'.join(str(station) for station in selected_stations)}_daily_df.csv",
                    ),
                    index=False,
                )

                # Save the annual weather data as csv without indexes
                pd.DataFrame(
                    {
                        "Annual_rf_mm": rain,
                        "Annual_ETo_Short_mm": eto_short,
                        "Annual_ETo_Tall_mm": eto_tall,
                    },
                    index=[0],
                ).to_csv(
                    os.path.join(
                        td,
                        f"{'+'.join(str(station) for station in selected_stations)}_annual_ave_df.csv",
                    ),
                    index=False,
                )

                pd.DataFrame(
                    {
                        "LongTerm_rf_mm": rain_long,
                        "LongTerm_ETo_Short_mm": eto_short_long,
                        "LongTerm_ETo_Tall_mm": eto_tall_long,
                    },
                    index=[0],
                ).to_csv(
                    os.path.join(
                        td,
                        f"{'+'.join(str(station) for station in selected_stations)}_longterm_ave_df.csv",
                    ),
                    index=False,
                )

                # Put everything into a zip file
                shutil.make_archive("Weather_data", "zip", os.path.join(td))

                zip_name = (
                    f"{'+'.join(str(num) for num in selected_stations)}"
                    + "_"
                    + str(dt.today().strftime("%d-%m-%Y"))
                )
                # Download the zip file
                with open("Weather_data.zip", "rb") as f:
                    st.download_button(
                        "Download weather data?", f, file_name=zip_name + ".zip"
                    )
        except ValueError:
            st.write("Haven't upload a bunch of shapefiles yet")

    if st.button("Start the extraction process", key="Extraction"):

        # A temporary output file
        with tempfile.TemporaryDirectory() as tmp_out:

            # Write out the general info
            follow_up(questionnaire_df, tmp_out)

            # Crop specific info
            land_management(questionnaire_df, crops, tmp_out)

            xlsx_path = glob.glob(pathname=os.path.join("input", "*.xlsx"))

            # Write into the inventory sheet
            wb = openpyxl.load_workbook(xlsx_path[0])

            # Fill in general info
            ws = wb["👤 General information"]

            # General information
            # Client name
            ws.cell(2, 2).value = questionnaire_df["client_name"].iloc[0]
            # Business name
            ws.cell(3, 2).value = questionnaire_df["business_name"].iloc[0]
            # Client email
            ws.cell(4, 2).value = questionnaire_df["email"].iloc[0]
            # Production year assessed
            ws.cell(5, 2).value = production_year

            # Location
            # Property name
            ws.cell(7, 2).value = questionnaire_df["property_name"].iloc[0]
            # Property address
            ws.cell(8, 2).value = questionnaire_df["property_address"].iloc[0]
            # State
            ws.cell(9, 2).value = STATE_REMAP.get(
                questionnaire_df["state"].iloc[0], questionnaire_df["state"].iloc[0]
            )
            # Farm map or paddock boundaries
            ws.cell(10, 2).value = questionnaire_df["upload_email_draw"].iloc[0]

            # Climate
            ## Rainfall & request ETo from DPIRD
            try:
                ws.cell(12, 2).value = questionnaire_df[
                    "property_av_annual_rainfall"
                ].iloc[0]
            except AttributeError:
                ws.cell(12, 2).value = "Didn't provide rainfall data"

            # Rainfall
            ws.cell(13, 2).value = rain
            # Evapotranspiration
            ws.cell(16, 2).value = eto_short
            ws.cell(16, 3).value = eto_tall

            # Software
            # Farm management software (Y/N)
            ws.cell(19, 2).value = questionnaire_df[
                "Do you use any Farm Management Practices software applications?"
            ].iloc[0]
            # List
            if isinstance(
                questionnaire_df["Please select the applications you use below"].iloc[
                    0
                ],
                str,
            ):
                ws.cell(20, 2).value = questionnaire_df["Please specify"].iloc[0]
            else:
                ws.cell(20, 2).value = (
                    questionnaire_df["Please select the applications you use below"]
                    .iloc[0]
                    .split(",")
                )
            # Practices
            # VRT
            strings = (
                questionnaire_df[
                    "Do you use variable rate technology (VRT) across your property ?"
                ]
                .iloc[0]
                .split("_")
            )
            ws.cell(22, 2).value = " ".join(strings)

            # Vegetation
            # Planting post_1990 (Y/N)
            try:
                ws.cell(26, 2).value = veg_df[" Location of plantings"].iloc[0]
                ws.cell(25, 2).value = "Y"
            except TypeError:
                ws.cell(25, 2).value = "N"
                ws.cell(26, 2).value = "N"

            # Electricity
            # Annual electricity use (KWh)
            ws.cell(28, 2).value = questionnaire_df[
                "What was your annual electricity consumption?"
            ].iloc[0]
            # Renewable (Y/N)
            ws.cell(29, 2).value = questionnaire_df[
                "Did you use renewable energy?"
            ].iloc[0]
            # Renewable source
            ws.cell(30, 2).value = questionnaire_df[
                "What was the source(s) of this renewable energy?"
            ].iloc[0]
            # % renewable
            ws.cell(31, 2).value = questionnaire_df[
                "What percentage of the total electricity consumption came from this source?"
            ].iloc[0]

            # Fuel
            fuels = ["diesel", "petrol", "LPG"]
            for i, fuel in enumerate(fuels):
                # Begining (L)
                ws.cell(33 + 3 * i, 2).value = questionnaire_df[
                    f"How much {fuel} did you have on hand at the start of the last calender year?"
                ].iloc[0]
                # Purchased (L)
                ws.cell(34 + 3 * i, 2).value = questionnaire_df[
                    f"How much {fuel} did you purchase throughout the year?"
                ].iloc[0]
                # End (l)
                ws.cell(35 + 3 * i, 2).value = questionnaire_df[
                    f"How much {fuel} did you have on hand at the end of the last calender year?"
                ].iloc[0]

            # Set the reference cell for offset below
            CropType_Header = Cell.Cell(ws, 46, 1)
            # Write into cells under corresponding crop types
            # using the refrence cell
            for i in range(12):  # Number of crop type
                croptype = CropType_Header.offset(i + 1)
                for crop in crops:
                    if crop in croptype.value.lower():
                        # Area sown
                        croptype.offset(column=1).value = questionnaire_df[
                            f"area_sown_{crop.lower()}"
                        ].iloc[0]
                        # Last year yield
                        croptype.offset(column=2).value = questionnaire_df[
                            f"av_yield_{crop.lower()}"
                        ].iloc[0]
                        # Burn (Y/N)
                        croptype.offset(column=5).value = questionnaire_df[
                            f"paddocks_burnt_{crop.lower()}"
                        ].iloc[0]
                        # Area burnt
                        if croptype.offset(column=5).value == "yes":
                            croptype.offset(column=6).value = (
                                questionnaire_df[f"windrow_burnt_{crop.lower()}"].iloc[
                                    0
                                ]
                                + questionnaire_df[f"area_burnt_{crop.lower()}"].iloc[0]
                            )  # Need update to specific crop type
                        else:
                            croptype.offset(column=6).value = 0

            # Fertiliser
            ws = wb["🛢️ Fertiliser Applied - Input"]
            # List of fertiliser applied breaks down by
            # crop type
            ferts = list_fert_chem(crop_specific_input, crops, questionnaire_df, "fert")
            write_products_to_ws(
                ws,
                crops,
                ferts,
                {1: "name", 2: "form", 4: _CROP_NAME, 6: "rate", 7: "area", 8: "times"},
            )

            # Chemical
            ws = wb["🧪 Chemical Applied - Input"]
            # List of chemical applied break downs
            # by crop
            chemicals = ["fungicide", "herbicide", "insecticide", "chem_other"]
            chems = {}
            for chem_type in chemicals:
                st.write(chem_type)
                chems[chem_type] = list_fert_chem(
                    crop_specific_input, crops, questionnaire_df, chem_type
                )
            for chem_type in chemicals:
                write_products_to_ws(
                    ws,
                    crops,
                    chems[chem_type],
                    {
                        1: "name",
                        2: "form",
                        16: _CROP_NAME,
                        17: "rate",
                        18: "area",
                        19: "times",
                    },
                )

            # Lime/gypsum
            ws = wb["🍋‍🟩Lime Product - Input"]
            # List of products (lime/dolomite and gypsum) applied
            # breaking down by crop type
            products_applied = to_soil_ame(questionnaire_df, crops)
            write_products_to_ws(
                ws,
                crops,
                products_applied,
                {1: "name", 2: "source", 4: _CROP_NAME, 5: "rate", 6: "area"},
            )

            #  Fuel usage - PW pathway will be in the future
            # ws = wb["⛽Fuel Usage - Input"]

            # Vegetation
            ws = wb["🌿 Vegetation - Input"]
            # A dictionary of vegetation planted
            vegetation = to_veg(veg_df, planting_shapes)
            # Write into the worksheet
            try:
                for i in range(len(vegetation)):
                    # Region
                    ws.cell(2 + i, 1).value = vegetation[i]["region"]
                    # Species
                    ws.cell(2 + i, 2).value = vegetation[i]["species"]
                    # Soil
                    ws.cell(2 + i, 4).value = vegetation[i]["soil"]
                    # Area
                    ws.cell(2 + i, 5).value = vegetation[i]["area"]
                    # Planted year
                    ws.cell(2 + i, 6).value = vegetation[i]["planted_year"]
                    # Age
                    ws.cell(2 + i, 7).value = vegetation[i]["age"]
            except TypeError:
                ws.cell(2, 1).value = "No planting"
                ws.cell(2, 2).value = "No planting"
                ws.cell(2, 4).value = "No planting"
                ws.cell(2, 5).value = "No planting"
                ws.cell(2, 6).value = "No planting"
                ws.cell(2, 7).value = "No planting"

            # Save the workbook
            wb.save(os.path.join(tmp_out, "Inventory_Sheet.xlsx"))

            # Create a zip to save follow ups question
            # and workbook
            shutil.make_archive("Question_Extract", "zip", tmp_out)

            # Name the file by the first property name
            zip_name = (
                str(questionnaire_df.loc[0, "property_name"])
                + "_"
                + str(dt.today().strftime("%d-%m-%Y"))
            )

            with open("Question_Extract.zip", "rb") as f:
                st.download_button(
                    "Download the extracted info", f, file_name=zip_name + ".zip"
                )

        files = [
            os.path.join(cwd, "Question_Extract.zip"),
            os.path.join(cwd, "Weather_data.zip"),
        ]
        remove_files(files)
else:
    st.header("Send to AIA")

    st.subheader("Disclaimer")
    st.write(
        "Before uploading the excel file, please open and save it so the data can be\nupdated accordingly"
    )

    ex_file = st.file_uploader("Upload your inventory sheet:", "xlsx")

    Crop = []
    desired_crop = []
    try:
        # Create a df using function
        df = to_data_frame(ex_file)
        df["Area sown (ha)"] = df["Area sown (ha)"].apply(lambda x: float(x))

        # Separate it by crop type
        Crop = by_crop_type(df)
        # Display the dataframe for checking
        if st.toggle("Do you want to check your input data frame?"):
            st.dataframe(Crop, hide_index=True)
            st.write("If there are no data, please refer to the text above")
        # Choose the desired crop to send a request
        desired_crop = st.multiselect(
            "Choose which crop to send your request:",
            df["Crop type"].loc[df["Area sown (ha)"] > 0].to_list(),
        )
    except TypeError:
        st.write("Haven't uploaded an inventory sheet yet")

    # Name the file by the first property name
    filename = st.text_input("Save the file as:", key="GAFF_file")

    if st.button("Run", key="AIA_API"):

        loc, rain_over = gen_info(ex_file)
        selected_indices = find_selected_indices(desired_crop, Crop)
        payload = build_aia_payload(Crop, selected_indices, rain_over)
        response = call_aia_api(payload)

        st.write(response.status_code)
        if response.status_code != 200:
            st.write(response.json())
        else:
            json_str = json.dumps(response.json(), indent=4)
            st.download_button(
                "Download the result from AIA's API",
                data=json_str.encode("utf-8"),
                file_name=filename + "_" + "_".join(desired_crop) + ".json",
                mime="application/json",
            )
