# Import the required function
import openpyxl


# Get the general information
def gen_info(excel_wb: str):
    # Read in the excel wb
    wb = openpyxl.load_workbook(excel_wb, data_only=True)
    # Activate the sheet
    ws = wb["👤 General information"]
    # Get the state loc info (assuming the cell does not change)
    loc = ws.cell(9, 2).value
    # Rainfall > 600mm
    if ws.cell(17, 2).value == "N":
        rain_over = False
    else:
        rain_over = True
    return loc, rain_over
